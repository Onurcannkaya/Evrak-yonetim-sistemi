"""
Evrak Yönetim Sistemi — Gelişmiş Araçlar (tools.py)
BatchProcessor, PDFMerger, ExcelExporter
"""
import os
import logging
from typing import List, Dict, Any

import fitz  # PyMuPDF
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from PyQt6.QtCore import QThread, pyqtSignal

logger = logging.getLogger("Tools")


# ══════════════════════ TOPLU İŞLEME ══════════════════════
class BatchWorker(QThread):
    """
    Bir dosya listesini sırayla Gemini'ye gönderir.
    Sinyaller:
        progress(int, int, str) → (mevcut_idx, toplam, dosya_adı)
        item_done(int, dict)    → (idx, sonuç_dict)
        all_done(list)          → tüm sonuçlar
        error(int, str)         → (idx, hata_mesajı)
    """
    progress = pyqtSignal(int, int, str)
    item_done = pyqtSignal(int, dict)
    all_done = pyqtSignal(list)
    error = pyqtSignal(int, str)

    def __init__(self, file_list: List[str], analyzer):
        super().__init__()
        self.file_list = file_list
        self.analyzer = analyzer
        self._cancelled = False

    def cancel(self):
        self._cancelled = True

    def run(self):
        results = []
        total = len(self.file_list)
        for i, fp in enumerate(self.file_list):
            if self._cancelled:
                break
            self.progress.emit(i + 1, total, os.path.basename(fp))
            try:
                res = self.analyzer.analyze_document(fp)
                res["_source_file"] = fp
                if "error" in res and res["error"]:
                    self.error.emit(i, res["error"])
                else:
                    self.item_done.emit(i, res)
                results.append(res)
            except Exception as e:
                self.error.emit(i, str(e))
                results.append({"error": str(e), "_source_file": fp})
        self.all_done.emit(results)


# ══════════════════════ PDF BİRLEŞTİRME ══════════════════════
class PDFMerger:
    """Birden fazla dosyayı tek bir PDF'e birleştirir."""

    @staticmethod
    def merge(file_paths: List[str], output_path: str) -> str:
        """
        Dosya listesini tek PDF olarak kaydet.
        PDF dosyaları doğrudan eklenir, resimler sayfaya dönüştürülür.
        """
        merged = fitz.open()

        for fp in file_paths:
            try:
                if fp.lower().endswith(".pdf"):
                    doc = fitz.open(fp)
                    merged.insert_pdf(doc)
                    doc.close()
                else:
                    # Resmi sayfaya dönüştür
                    img = fitz.open(fp)
                    rect = img[0].rect
                    pdf_bytes = img.convert_to_pdf()
                    img.close()
                    img_pdf = fitz.open("pdf", pdf_bytes)
                    merged.insert_pdf(img_pdf)
                    img_pdf.close()
            except Exception as e:
                logger.error(f"PDF birleştirme hatası ({fp}): {e}")

        merged.save(output_path)
        merged.close()
        return output_path


# ══════════════════════ EXCEL AKTARIMI ══════════════════════
class ExcelExporter:
    """Veritabanı verilerini .xlsx dosyasına aktarır."""

    HEADER_FILL = PatternFill(start_color="4E9A06", end_color="4E9A06", fill_type="solid")
    HEADER_FONT = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    CELL_FONT = Font(name="Segoe UI", size=10)

    @staticmethod
    def export(data: List[Dict[str, Any]], output_path: str, 
               columns: List[str] = None) -> str:
        """
        Veri listesini Excel dosyasına aktarır.
        
        Args:
            data: Dışa aktarılacak kayıt listesi
            output_path: .xlsx dosya yolu
            columns: Sütun key listesi (None ise varsayılan kullanılır)
        """
        if columns is None:
            columns = ["id", "mahalle", "ada", "parsel", "tarih", "raw_text"]
        
        # Türkçe başlıklar
        header_map = {
            "id": "No",
            "mahalle": "Mahalle",
            "p_mahalle": "Mahalle",
            "ada": "Ada",
            "parsel": "Parsel",
            "tarih": "Tarih",
            "extracted_date": "Tarih",
            "raw_text": "Okunan Metin",
            "file_path": "Dosya Yolu",
            "doc_type": "Belge Türü",
        }

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Evrak Raporu"

        # Başlık satırı
        for col_idx, key in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header_map.get(key, key))
            cell.fill = ExcelExporter.HEADER_FILL
            cell.font = ExcelExporter.HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        # Veri satırları
        for row_idx, record in enumerate(data, 2):
            for col_idx, key in enumerate(columns, 1):
                val = record.get(key, "")
                if val is None:
                    val = ""
                cell = ws.cell(row=row_idx, column=col_idx, value=str(val))
                cell.font = ExcelExporter.CELL_FONT

        # Sütun genişliklerini otomatik ayarla
        for col_idx, key in enumerate(columns, 1):
            max_len = len(header_map.get(key, key))
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, min(len(str(cell.value)), 50))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_len + 4

        wb.save(output_path)
        return output_path
